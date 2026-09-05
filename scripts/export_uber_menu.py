#!/usr/bin/env python3
"""Bambu Menu → src/data/uber-menu.generated.ts

Source of truth: `_Bambu _ Uber Menu  (3).xlsx` in repo root (confirmed menu data).
- One sheet per category (`1.Bambu Special Menu`, `2.Sweet Dessert`, …).
  Only the sheets listed in SECTION_SPECS are read — the workbook also holds
  a consolidated `uber-menu-prod` sheet and legacy per-store food sheets
  (CABRA/CANLEY) whose Product-ID prefixes clash with the current ones.
- Each row carries a `Product ID` like `P0101` → PXXYY where XX = category,
  YY = item; rows are grouped into one TS const per sheet.
- Rows flagged in the `Inactive` column are skipped (hidden from the site).
- Images come from the `STT Ảnh` / image-link column (Google Drive share
  URLs); the app's product-image resolver rewrites them to embeddable
  thumbnail URLs.

Commands:
  python3 scripts/export_uber_menu.py
      Reads the Uber Menu workbook → writes TS. (No-op if the workbook is missing.)

Install: pip install openpyxl
"""
from __future__ import annotations

import re
import sys
import unicodedata
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
XLSX = ROOT / "_Bambu _ Uber Menu  (3).xlsx"
OUT = ROOT / "src" / "data" / "uber-menu.generated.ts"

# Worksheet name → the TS const it is emitted as.
# Order here is the order consts appear in the TS file.
SECTION_SPECS = [
    {"sheet": "1.Bambu Special Menu", "const": "BAMBU_SPECIAL"},
    {"sheet": "2.Sweet Dessert", "const": "SWEET_DESSERT"},
    {"sheet": "3.Smashed fruit & Sweets", "const": "SMASHED_FRUIT"},
    {"sheet": "4.Matcha Drinks", "const": "MATCHA"},
    {"sheet": "5.Fruit Drinks - Tea", "const": "FRUIT_DRINKS_TEA"},
    {"sheet": "6.Fresh Juice", "const": "FRESH_JUICE"},
    {"sheet": "7.Over Ice", "const": "OVER_ICE"},
    {"sheet": "8.Espresso (Hot)", "const": "ESPRESSO_HOT"},
    {"sheet": "9.Ice Blended", "const": "ICE_BLENDED"},
    {"sheet": "10.Smoothies", "const": "SMOOTHIES"},
    {"sheet": "11. Foods", "const": "FOODS"},
    {"sheet": "12. Breakfast - Cabra store onl", "const": "BREAKFAST_CABRA"},
    {"sheet": "13. Whats new", "const": "WHATS_NEW"},
]

ALL_CONSTS = [s["const"] for s in SECTION_SPECS]
SHEET_TO_CONST = {s["sheet"]: s["const"] for s in SECTION_SPECS}

# Header label (normalised) → record field. The category sheets label their
# columns in Vietnamese; older English labels are kept as fallbacks.
COLUMN_MAP = {
    "stt": "stt",
    "no.": "stt",
    "product id": "productId",
    "tên tiếng việt": "nameVi",
    "vietnamese name": "nameVi",
    "tên tiếng anh": "nameEn",
    "english name": "nameEn",
    "tên uber": "nameUber",
    "name": "nameUber",
    "inactive": "inactive",
    "type": "type",
    "size": "size",
    "tag": "tag",
    "quantity": "quantity",
    "giá tiền offline store = pickup (aud)": "pricePickup",
    "pickup price (aud)": "pricePickup",
    "giá tiền uber (aud)": "priceUber",
    "delivery price (aud)": "priceUber",
    "mô tả": "description",
    "description": "description",
    "stt ảnh": "photoStt",
    "stt anh": "photoStt",
    "image link": "photoStt",
    "ghi chú": "notes",
    "notes": "notes",
}


def _norm_key(s) -> str:
    s = unicodedata.normalize("NFC", str(s))
    return re.sub(r"\s+", " ", s.strip()).lower()


def ts_str(s: str | None) -> str:
    if s is None:
        return "null"
    s = str(s).replace("\\", "\\\\").replace("'", "\\'")
    s = s.replace("\r\n", " ").replace("\n", " ").replace("\r", " ")
    s = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", s)
    return "'" + s + "'"


def ts_num(n: float | int | None) -> str:
    if n is None:
        return "null"
    if isinstance(n, float) and n == int(n):
        return str(int(n))
    return str(round(n, 2))


def cell_str(v) -> str | None:
    if v is None:
        return None
    if isinstance(v, float) and v == int(v):
        v = int(v)
    s = str(v).strip()
    return s if s else None


def cell_float(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip().replace(",", ""))
    except (TypeError, ValueError):
        return None


def cell_stt(v) -> int | None:
    if v is None:
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _is_vietnamese(s: str) -> bool:
    """True if the text carries a Vietnamese-only mark: đ/Đ or a toned vowel."""
    for ch in s:
        if ch in "đĐ":
            return True
        nfd = unicodedata.normalize("NFD", ch)
        if nfd[0].isascii() and any(unicodedata.category(m) == "Mn" for m in nfd[1:]):
            return True
    return False


def _strip_diacritics(s: str) -> str:
    """Vietnamese text folded to plain ASCII letters: 'Bò bía ngọt' → 'Bo bia ngot'."""
    s = s.replace("đ", "d").replace("Đ", "D")
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def orient_names(name_vi: str, name_en: str) -> tuple[str, str]:
    """Some workbook rows have the Vietnamese and English cells filled in swapped.

    Diacritics arbitrate: reorder only when exactly one side is unambiguously
    Vietnamese and it is sitting in the English column. Rows where both or
    neither side carries diacritics are left as authored — they cannot be told
    apart without a human reading them.

    A row whose two cells hold the same Vietnamese text has no English name at
    all; the site is English-first, so it shows that name unaccented.
    """
    if _is_vietnamese(name_en) and not _is_vietnamese(name_vi):
        name_vi, name_en = name_en, name_vi
    if name_en == name_vi and _is_vietnamese(name_en):
        name_en = _strip_diacritics(name_en)
    return name_vi, name_en


def find_header_row(ws) -> tuple[int, dict[str, int]] | None:
    """Locate the row that holds the column headers and map field → col index."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if not row:
            continue
        norm = {_norm_key(c): j for j, c in enumerate(row) if isinstance(c, str)}
        if "product id" in norm and ("tên tiếng anh" in norm or "english name" in norm or "tên tiếng việt" in norm):
            col: dict[str, int] = {}
            for label, field in COLUMN_MAP.items():
                if label in norm:
                    col[field] = norm[label]
            return i, col
    return None


def apply_photo_hyperlink(rec: dict, ws, *, excel_row: int, photo_col_0based: int) -> None:
    """Prefer a cell hyperlink target over its display text for the image link."""
    c = ws.cell(row=excel_row, column=photo_col_0based + 1)
    hyp = getattr(c, "hyperlink", None)
    target = getattr(hyp, "target", None) if hyp is not None else None
    if target:
        t = str(target).strip()
        if t.lower().startswith(("http://", "https://")):
            rec["photoStt"] = t


def parse_workbook(path: Path) -> dict[str, list[dict]]:
    wb = openpyxl.load_workbook(path, data_only=True)
    buckets: dict[str, list[dict]] = {c: [] for c in ALL_CONSTS}
    try:
        for spec in SECTION_SPECS:
            sheet, const = spec["sheet"], spec["const"]
            if sheet not in wb.sheetnames:
                print(f"Warning: sheet {sheet!r} missing from workbook, skipped", file=sys.stderr)
                continue
            ws = wb[sheet]
            found = find_header_row(ws)
            if found is None:
                print(f"Warning: no header row found in sheet {sheet!r}, skipped", file=sys.stderr)
                continue
            header_idx, col = found

            def get(cells, field):
                idx = col.get(field)
                return cells[idx] if idx is not None and idx < len(cells) else None

            for r_i, row in enumerate(
                ws.iter_rows(min_row=header_idx + 2, values_only=True),
                start=header_idx + 2,
            ):
                cells = list(row)
                pid = cell_str(get(cells, "productId"))
                name_en = cell_str(get(cells, "nameEn")) or ""
                name_vi = cell_str(get(cells, "nameVi")) or ""
                if not pid or (not name_en and not name_vi):
                    continue
                name_vi, name_en = orient_names(name_vi, name_en)
                # Skip rows flagged inactive (any non-empty value).
                if cell_str(get(cells, "inactive")):
                    continue
                rec = {
                    "productId": pid,
                    "stt": cell_stt(get(cells, "stt")),
                    "nameVi": name_vi,
                    "nameEn": name_en,
                    "nameUber": cell_str(get(cells, "nameUber")),
                    "type": cell_str(get(cells, "type")),
                    "size": cell_str(get(cells, "size")),
                    "tag": cell_str(get(cells, "tag")),
                    "pricePickup": cell_float(get(cells, "pricePickup")),
                    "priceUber": cell_float(get(cells, "priceUber")),
                    "description": cell_str(get(cells, "description")),
                    "photoStt": cell_str(get(cells, "photoStt")),
                    "notes": cell_str(get(cells, "notes")),
                }
                if "photoStt" in col:
                    apply_photo_hyperlink(rec, ws, excel_row=r_i, photo_col_0based=col["photoStt"])
                buckets[const].append(rec)
    finally:
        wb.close()
    return buckets


def emit_ts(buckets: dict[str, list[dict]], *, source_note: str) -> str:
    def emit_list(name: str, rows: list[dict]) -> str:
        lines = [f"export const {name}: UberMenuRow[] = ["]
        for r in rows:
            photo = r.get("photoStt")
            photo_ts = ts_str(photo) if isinstance(photo, str) else "null"
            lines.append(
                "  {"
                + f"productId: {ts_str(r.get('productId'))}, "
                + f"stt: {ts_num(r.get('stt'))}, "
                + f"nameVi: {ts_str(r.get('nameVi'))}, "
                + f"nameEn: {ts_str(r.get('nameEn'))}, "
                + f"nameUber: {ts_str(r.get('nameUber')) if r.get('nameUber') else 'null'}, "
                + f"type: {ts_str(r.get('type')) if r.get('type') else 'null'}, "
                + f"size: {ts_str(r.get('size')) if r.get('size') else 'null'}, "
                + f"tag: {ts_str(r.get('tag')) if r.get('tag') else 'null'}, "
                + f"pricePickup: {ts_num(r.get('pricePickup'))}, "
                + f"priceUber: {ts_num(r.get('priceUber'))}, "
                + f"description: {ts_str(r.get('description')) if r.get('description') else 'null'}, "
                + f"photoStt: {photo_ts}, "
                + f"notes: {ts_str(r.get('notes')) if r.get('notes') else 'null'}, "
                + "},"
            )
        lines.append("];")
        return "\n".join(lines)

    header = f"""/**
 * AUTO-GENERATED from `{source_note}`
 * Do not edit by hand — run: `bun run export-menu` (or `python3 scripts/export_uber_menu.py`)
 *
 * Rows are grouped into consts by their Product ID prefix (P01…P12).
 * Rows flagged `Inactive` in the workbook are omitted.
 */

export type UberMenuRow = {{
  productId: string | null;
  stt: number | null;
  nameVi: string;
  nameEn: string;
  nameUber: string | null;
  type: string | null;
  size: string | null;
  tag: string | null;
  pricePickup: number | null;
  priceUber: number | null;
  description: string | null;
  photoStt: number | string | null;
  notes: string | null;
}};

export type IcedCoffeeExtra = {{
  nameVi: string;
  nameEn: string;
  priceLine: string;
}};

"""

    parts = [header]
    for const in ALL_CONSTS:
        parts.append(emit_list(const, buckets.get(const, [])))
        parts.append("\n")
    # Kept for backward compatibility with any importers of the extras list.
    parts.append("export const ICED_COFFEE_EXTRAS: IcedCoffeeExtra[] = [];")
    parts.append("\n")
    return "\n".join(parts)


def main() -> None:
    if not XLSX.exists():
        print(f"No {XLSX.name} — skipped export (keeping {OUT.name}).")
        return
    buckets = parse_workbook(XLSX)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(emit_ts(buckets, source_note=XLSX.name), encoding="utf-8")
    counts = ", ".join(f"{len(buckets.get(c, []))} {c}" for c in ALL_CONSTS)
    print(f"Wrote {OUT} from {XLSX.name}\n  {counts}")


if __name__ == "__main__":
    main()
